# utils/excel.py
import re
import tempfile
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from typing import List, Dict, Optional


# ======================
# ЭКСПОРТ (для админа)
# ======================

def export_tasks_to_excel(tasks) -> str:
    """Генерирует Excel-файл со всеми задачами и возвращает путь к нему."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Все задачи"
    headers = [
        "practice_name", "start_date", "end_date", "full_name",
        "tg_username", "phone", "task_description", "status", "next_reminder"
    ]
    ws.append(headers)
    for task in tasks:
        row = [
            task["practice_name"],
            task["start_date"] or "",
            task["end_date"],
            task["full_name"],
            task["tg_username"] or "—",
            task["phone"] or "—",
            task["task_description"],
            task["status"],
            task["next_reminder"] or ""
        ]
        ws.append(row)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        return tmp.name


# ======================
# ИМПОРТ (парсинг Excel)
# ======================

def parse_excel(file_path: str) -> List[Dict]:
    """Парсит Excel и возвращает список задач в унифицированном формате."""
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell).strip().lower() if cell else "" for cell in rows[0]]
    tasks = []

    # Проверяем, есть ли обязательные колонки
    if "full_name" not in headers or "end_date" not in headers:
        raise ValueError("Файл должен содержать колонки: full_name и end_date")

    # Индексы колонок
    full_name_idx = headers.index("full_name")
    end_date_idx = headers.index("end_date")

    practice_name_idx = headers.index("practice_name") if "practice_name" in headers else None
    task_desc_idx = headers.index("task_description") if "task_description" in headers else None
    start_date_idx = headers.index("start_date") if "start_date" in headers else None
    tg_username_idx = headers.index("tg_username") if "tg_username" in headers else None
    phone_idx = headers.index("phone") if "phone" in headers else None

    for row in rows[1:]:
        if not row or not row[full_name_idx]:
            continue

        full_name = str(row[full_name_idx]).strip()
        end_date_raw = row[end_date_idx]
        end_date = normalize_date(end_date_raw)
        if not end_date:
            continue

        # Определяем описание
        if task_desc_idx is not None and row[task_desc_idx]:
            description = str(row[task_desc_idx]).strip()
        elif practice_name_idx is not None and row[practice_name_idx]:
            description = str(row[practice_name_idx]).strip()
        else:
            description = "Без описания"

        practice_name = None
        if practice_name_idx is not None and row[practice_name_idx]:
            practice_name = str(row[practice_name_idx]).strip()

        start_date = normalize_date(row[start_date_idx]) if start_date_idx and row[start_date_idx] else None

        tg_username = None
        if tg_username_idx and row[tg_username_idx]:
            un = str(row[tg_username_idx]).strip()
            tg_username = un if un.startswith("@") else f"@{un}"

        phone = str(row[phone_idx]).strip() if phone_idx and row[phone_idx] else None

        tasks.append({
            "full_name": full_name,
            "practice_name": practice_name,
            "start_date": start_date,
            "end_date": end_date,
            "description": description,
            "tg_username": tg_username,
            "phone": phone
        })

    return tasks


def normalize_date(date_value) -> Optional[str]:
    """Приводит дату к формату YYYY-MM-DD."""
    if not date_value:
        return None
    if isinstance(date_value, datetime):
        return date_value.strftime("%Y-%m-%d")
    if isinstance(date_value, str):
        date_str = date_value.strip()
        # ДД.ММ.ГГГГ
        if re.match(r"\d{1,2}\.\d{1,2}\.\d{4}", date_str):
            try:
                dt = datetime.strptime(date_str, "%d.%m.%Y")
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                pass
        # ГГГГ-ММ-ДД
        elif re.match(r"\d{4}-\d{2}-\d{2}", date_str):
            return date_str
    return None
def create_excel_template() -> str:
    """Создаёт Excel-файл с двумя шаблонами: полным и кратким."""
    from openpyxl import Workbook

    wb = Workbook()

    # Лист 1: Полный формат
    ws1 = wb.active
    ws1.title = "Полный формат"
    ws1.append(["practice_name", "start_date", "end_date", "full_name", "tg_username", "phone"])
    ws1.append(["Производственная практика", "2025-06-01", "2025-07-15", "Иванов И.И.", "@ivanov", "+79991234567"])

    # Лист 2: Краткий формат
    ws2 = wb.create_sheet("Краткий формат")
    ws2.append(["full_name", "task_description", "end_date"])
    ws2.append(["Петров П.П.", "Подготовить программу практики", "15.07.2025"])

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        return tmp.name


from io import BytesIO
from openpyxl import load_workbook


def parse_excel_from_bytes(file_bytes: BytesIO) -> List[Dict]:
    """Парсит Excel из BytesIO (без сохранения на диск)."""
    wb = load_workbook(filename=file_bytes, read_only=True, data_only=True)
    ws = wb.active
    # ... остальной код как в parse_excel(), начиная с rows = list(ws.iter_rows(...))
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell).strip().lower() if cell else "" for cell in rows[0]]
    tasks = []

    if "full_name" not in headers or "end_date" not in headers:
        raise ValueError("Файл должен содержать колонки: full_name и end_date")

    full_name_idx = headers.index("full_name")
    end_date_idx = headers.index("end_date")
    practice_name_idx = headers.index("practice_name") if "practice_name" in headers else None
    task_desc_idx = headers.index("task_description") if "task_description" in headers else None
    start_date_idx = headers.index("start_date") if "start_date" in headers else None
    tg_username_idx = headers.index("tg_username") if "tg_username" in headers else None
    phone_idx = headers.index("phone") if "phone" in headers else None

    for row in rows[1:]:
        if not row or not row[full_name_idx]:
            continue

        full_name = str(row[full_name_idx]).strip()
        end_date_raw = row[end_date_idx]
        end_date = normalize_date(end_date_raw)
        if not end_date:
            continue

        if task_desc_idx is not None and row[task_desc_idx]:
            description = str(row[task_desc_idx]).strip()
        elif practice_name_idx is not None and row[practice_name_idx]:
            description = str(row[practice_name_idx]).strip()
        else:
            description = "Без описания"

        practice_name = None
        if practice_name_idx is not None and row[practice_name_idx]:
            practice_name = str(row[practice_name_idx]).strip()

        start_date = normalize_date(row[start_date_idx]) if start_date_idx and row[start_date_idx] else None
        tg_username = None
        if tg_username_idx and row[tg_username_idx]:
            un = str(row[tg_username_idx]).strip()
            tg_username = un if un.startswith("@") else f"@{un}"
        phone = str(row[phone_idx]).strip() if phone_idx and row[phone_idx] else None

        tasks.append({
            "full_name": full_name,
            "practice_name": practice_name,
            "start_date": start_date,
            "end_date": end_date,
            "description": description,
            "tg_username": tg_username,
            "phone": phone
        })

    return tasks